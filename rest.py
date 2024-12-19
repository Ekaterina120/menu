import tkinter as tk
from tkinter import ttk
import openpyxl
import sqlite3
from tkinter import Tk, Toplevel, Frame, Label, Button, Entry, messagebox, Canvas, Scrollbar, BooleanVar, Checkbutton

# Подключение к базе данных
def connect_db():
    conn = sqlite3.connect('menu.db')
    return conn


# Загрузка данных из базы данных
def fetch_data(query, params=None):
    conn = sqlite3.connect('menu.db')
    cursor = conn.cursor()
    cursor.execute(query, params or ())
    result = cursor.fetchall()
    conn.close()
    return result


# Выполнение изменений в базе данных
def execute_query(query, params=()):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute(query, params)
    conn.commit()
    conn.close()


# Пароли для администратора и повара
USER_CREDENTIALS = {
    "admin": "admin123",  # Пароль для администратора
    "chef": "chef123"  # Пароль для повара
}


def authenticate_user(role, password):
    return USER_CREDENTIALS.get(role) == password


def login(role, callback):
    def attempt_login():
        password = password_entry.get()

        if authenticate_user(role, password):
            login_window.destroy()
            callback()
        else:
            messagebox.showerror("Ошибка", "Неверный пароль.")

    login_window = tk.Toplevel(root)
    login_window.title(f"Вход для {role.capitalize()}")
    login_window.geometry("300x150+600+300")

    ttk.Label(login_window, text="Введите пароль:", font=("Arial", 12)).pack(pady=10)
    password_entry = ttk.Entry(login_window, font=("Arial", 12), show="*")
    password_entry.pack(pady=5)

    ttk.Button(login_window, text="Войти", command=attempt_login).pack(pady=10)


# Создание основного окна приложения
class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Система управления меню")
        self.root.geometry("400x300+600+300")
        style = ttk.Style()
        style.configure("TButton", font=("Arial", 12), padding=6)
        style.configure("TLabel", font=("Arial", 14))

        ttk.Label(self.root, text="Выберите роль:").pack(pady=20)
        ttk.Button(self.root, text="Администратор", width=20, command=self.open_admin).pack(pady=10)
        ttk.Button(self.root, text="Повар", width=20, command=self.open_chef).pack(pady=10)
        ttk.Button(self.root, text="Клиент", width=20, command=self.open_client).pack(pady=10)

    def open_admin(self):
        login("admin", lambda: AdminWindow(self.root))

    def open_chef(self):
        login("chef", lambda: ChefWindow(self.root))

    def open_client(self):
        ClientWindow(self.root)


# Класс окна повара
class ChefWindow:
    def __init__(self, parent):
        self.window = Toplevel(parent)
        self.window.title("Повар")
        self.window.state("zoomed")  # Увеличенный размер окна
        self.window.config(bg="#D3D3D3")  # Голубой фон

        # Рамка для меню (центр) с прокруткой
        menu_container = Frame(self.window, bg="#FFFFFF", bd=2, relief="groove")
        menu_container.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        # Создание Canvas для прокрутки
        menu_canvas = Canvas(menu_container, bg="#FFFFFF")
        menu_canvas.pack(side="left", fill="both", expand=True)

        # Вертикальная прокрутка
        menu_scrollbar = Scrollbar(menu_container, orient="vertical", command=menu_canvas.yview)
        menu_scrollbar.pack(side="right", fill="y")
        menu_canvas.configure(yscrollcommand=menu_scrollbar.set)

        # Рамка внутри Canvas для отображения меню
        self.menu_frame = Frame(menu_canvas, bg="#FFFFFF")
        menu_canvas.create_window((0, 0), window=self.menu_frame, anchor="nw")

        # Обновление размеров Canvas при изменении содержимого
        def on_frame_configure(event):
            menu_canvas.configure(scrollregion=menu_canvas.bbox("all"))

        self.menu_frame.bind("<Configure>", on_frame_configure)

        # Рамка для стоп-листа (справа)
        self.stopped_frame = Frame(self.window, bg="#FFFFFF", bd=2, relief="groove", width=250)
        self.stopped_frame.pack(side="left", fill="y", padx=10, pady=10)

        # Загрузка меню и стоп-листа
        self.load_menu()
        self.load_stopped_dishes()

        # Кнопка для добавления блюда
        Button(
            self.window,
            text="Добавить блюдо",
            font=("Arial", 14),
            bg="#EEE8CD",
            fg="#333",
            command=self.add_new_dish,
        ).pack(pady=20)

    def add_new_dish(self):
        """Открывает окно для добавления нового блюда."""
        add_window = Toplevel(self.window)
        add_window.title("Добавить новое блюдо")
        add_window.geometry("600x700")
        add_window.config(bg="#FFFFFF")

        Label(
            add_window,
            text="Добавить новое блюдо",
            font=("Arial", 16, "bold"),
            bg="#FFFFFF",
            fg="#333",
        ).pack(pady=10)

        # Поля ввода для названия и цены блюда
        Label(add_window, text="Название блюда:", font=("Arial", 12), bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10,
                                                                                                    pady=5)
        dish_name_entry = Entry(add_window, font=("Arial", 12), bg="#F5F5F5")
        dish_name_entry.pack(fill="x", padx=10)

        Label(add_window, text="Цена блюда (₽):", font=("Arial", 12), bg="#FFFFFF", fg="#333").pack(anchor="w", padx=10,
                                                                                                    pady=5)
        dish_price_entry = Entry(add_window, font=("Arial", 12), bg="#F5F5F5")
        dish_price_entry.pack(fill="x", padx=10)

        # Контейнер с прокруткой для списка ингредиентов
        Label(add_window, text="Выберите ингредиенты:", font=("Arial", 12), bg="#FFFFFF", fg="#333").pack(anchor="w",
                                                                                                          padx=10,
                                                                                                          pady=10)

        scroll_frame = Frame(add_window, bg="#FFFFFF")
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=5)

        canvas = Canvas(scroll_frame, bg="#FFFFFF")
        canvas.pack(side="left", fill="both", expand=True)

        scrollbar = Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")

        ingredient_frame = Frame(canvas, bg="#FFFFFF")
        canvas.create_window((0, 0), window=ingredient_frame, anchor="nw")

        canvas.configure(yscrollcommand=scrollbar.set)

        # Функция для обновления размеров canvas
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        ingredient_frame.bind("<Configure>", on_frame_configure)

        # Получение списка ингредиентов из базы данных
        ingredients = fetch_data("SELECT ingredient_id, ingredient_name, quantity FROM ingredient")

        ingredient_checkboxes = {}

        for ingredient in ingredients:
            row = Frame(ingredient_frame, bg="#FFFFFF")
            row.pack(fill="x", pady=2)

            # Чекбокс для выбора ингредиента
            selected = BooleanVar()
            ingredient_checkboxes[ingredient[0]] = selected
            Checkbutton(
                row,
                text=f"{ingredient[1]} (Доступно: {ingredient[2]})",
                variable=selected,
                bg="#FFFFFF",
                anchor="w",
            ).pack(side="left", padx=5)

        # Кнопка сохранения блюда
        Button(
            add_window,
            text="Сохранить блюдо",
            font=("Arial", 12),
            bg="#EEE8CD",
            fg="#333",
            command=lambda: self.save_new_dish(dish_name_entry, dish_price_entry, ingredient_checkboxes, add_window),
        ).pack(pady=10)

    def save_new_dish(self, name_entry, price_entry, checkboxes, window):
        """Сохраняет новое блюдо и его ингредиенты в базе данных."""
        # Получение данных блюда
        dish_name = name_entry.get().strip()
        dish_price = price_entry.get().strip()

        if not dish_name:
            messagebox.showerror("Ошибка", "Название блюда не может быть пустым.")
            return

        if not dish_price.isdigit():
            messagebox.showerror("Ошибка", "Цена должна быть числом.")
            return

        dish_price = int(dish_price)

        # Сохранение блюда в базе данных
        try:
            # Генерация нового dish_id
            new_dish_id = fetch_data("SELECT IFNULL(MAX(dish_id), 0) + 1 FROM dish")[0][0]

            # Добавление нового блюда в таблицу dish
            execute_query(
                "INSERT INTO dish (dish_id, dish_name, price) VALUES (?, ?, ?)",
                (new_dish_id, dish_name, dish_price),
            )

            # Сохранение связей между блюдом и ингредиентами
            for ingredient_id, selected in checkboxes.items():
                if selected.get():  # Проверяем, выбран ли ингредиент
                    # Проверяем, достаточно ли ингредиента на складе
                    available_quantity = fetch_data(
                        "SELECT quantity FROM ingredient WHERE ingredient_id = ?",
                        (ingredient_id,)
                    )[0][0]

                    if available_quantity <= 0:
                        messagebox.showerror(
                            "Ошибка",
                            f"Ингредиента '{ingredient_id}' недостаточно на складе."
                        )
                        return

                    # Сохраняем связь между блюдом и ингредиентом
                    execute_query(
                        "INSERT INTO dishingredient (dish_id, ingredient_id) VALUES (?, ?)",
                        (new_dish_id, ingredient_id),
                    )

            # Успешное сохранение
            messagebox.showinfo("Успех", f"Блюдо '{dish_name}' успешно добавлено в меню.")
            window.destroy()

            # Обновление меню
            self.load_menu()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при сохранении блюда: {str(e)}")

    def load_menu(self):
        for widget in self.menu_frame.winfo_children():
            widget.destroy()

        dishes = fetch_data("SELECT d.dish_id, d.dish_name, d.price FROM dish d")
        Label(self.menu_frame, text="Меню:", font=("Arial", 16, "bold"), bg="#FFFFFF", fg="#333").pack(anchor="w",
                                                                                                       padx=10, pady=5)

        for dish in dishes:
            frame = Frame(self.menu_frame, bg="#FFFFFF")
            frame.pack(fill="x", padx=10, pady=5)

            Label(
                frame,
                text=f"{dish[1]} - {dish[2]} ₽",
                font=("Arial", 12),
                bg="#FFFFFF",
                anchor="w",
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Проверить ингредиенты",
                font=("Arial", 10),
                bg="#FFF8DC",
                fg="#333",
                command=lambda r=dish: self.check_ingredients(r),
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Добавить в стоп-лист",
                font=("Arial", 10),
                bg="#EEE8CD",
                fg="#333",
                command=lambda r=dish: self.add_to_stoplist(r),
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Управлять ингредиентами",
                font=("Arial", 10),
                bg="#CDC8B1",
                fg="#333",
                command=lambda r=dish: self.manage_ingredients(r),
            ).pack(side="left", padx=5)

            # Кнопка для удаления блюда
            Button(
                frame,
                text="Удалить",
                font=("Arial", 10),
                bg="#FFB6C1",
                fg="#333",
                command=lambda r=dish: self.delete_dish(r),
            ).pack(side="left", padx=5)

    def delete_dish(self, dish):
        """Удаляет блюдо из базы данных и обновляет меню."""
        try:
            # Удаление из таблицы блюд
            execute_query("DELETE FROM dish WHERE dish_id = ?", (dish[0],))

            # Также удаляем связи блюда с ингредиентами
            execute_query("DELETE FROM dishingredient WHERE dish_id = ?", (dish[0],))

            # Сообщение об успешном удалении
            self.show_message(f"Блюдо '{dish[1]}' успешно удалено из меню.")

            # Обновление интерфейса
            self.load_menu()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при удалении блюда: {str(e)}")

    def load_stopped_dishes(self):
        for widget in self.stopped_frame.winfo_children():
            widget.destroy()

        stopped_dishes = fetch_data("SELECT dish_id, dish_name, price FROM stopped_dishes")
        Label(self.stopped_frame, text="Стоп-лист:", font=("Arial", 16, "bold"), bg="#FFFFFF", fg="#333").pack(
            anchor="w", padx=10, pady=5
        )

        for dish in stopped_dishes:
            frame = Frame(self.stopped_frame, bg="#FFFFFF")
            frame.pack(fill="x", padx=10, pady=5)

            Label(
                frame,
                text=f"{dish[1]} - {dish[2]} ₽",
                font=("Arial", 12),
                bg="#FFFFFF",
                anchor="w",
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Вернуть в меню",
                font=("Arial", 10),
                bg="#EEE8CD",
                fg="#333",
                command=lambda r=dish: self.restore_dish(r),
            ).pack(side="left", padx=5)

    def check_ingredients(self, dish):
        ingredients = fetch_data(
            "SELECT i.ingredient_name, i.quantity FROM ingredient i "
            "JOIN dishingredient di ON i.ingredient_id = di.ingredient_id "
            "WHERE di.dish_id = ?",
            (dish[0],),
        )

        missing = [ing[0] for ing in ingredients if ing[1] <= 0]

        if missing:
            self.add_to_stoplist(dish, missing)
        else:
            self.show_message(f"Все ингредиенты для блюда '{dish[1]}' в наличии.")

    def add_to_stoplist(self, dish, missing=None, refresh_ui=True):
        """Добавляет блюдо в стоп-лист."""
        execute_query("DELETE FROM dish WHERE dish_id = ?", (dish[0],))
        execute_query(
            "INSERT INTO stopped_dishes (dish_id, dish_name, price) VALUES (?, ?, ?)",
            (dish[0], dish[1], dish[2]),
        )
        if missing:
            self.show_message(
                f"Блюдо '{dish[1]}' добавлено в стоп-лист. Отсутствуют ингредиенты: {', '.join(missing)}"
            )
        else:
            self.show_message(f"Блюдо '{dish[1]}' добавлено в стоп-лист вручную.")

        # Обновляем только стоп-лист и меню, если это нужно
        if refresh_ui:
            self.load_menu()
            self.load_stopped_dishes()

    def manage_ingredients(self, dish):
        ingredients = fetch_data(
            "SELECT i.ingredient_id, i.ingredient_name, i.quantity FROM ingredient i "
            "JOIN dishingredient di ON i.ingredient_id = di.ingredient_id "
            "WHERE di.dish_id = ?",
            (dish[0],),
        )

        manage_window = Toplevel(self.window)
        manage_window.title(f"Ингредиенты для блюда {dish[1]}")
        manage_window.geometry("500x600")
        manage_window.config(bg="#FFFFFF")

        Label(
            manage_window,
            text=f"Ингредиенты для блюда: {dish[1]}",
            font=("Arial", 14),
            bg="#FFFFFF",
            fg="#333",
        ).pack(pady=10)

        frame = Frame(manage_window, bg="#FFFFFF")
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Храним Entry виджеты для обновления количеств
        self.ingredient_entries = {}

        for ingredient in ingredients:
            row = Frame(frame, bg="#FFFFFF")
            row.pack(fill="x", padx=5, pady=5)

            # Отображение имени ингредиента
            Label(
                row,
                text=f"{ingredient[1]}:",
                font=("Arial", 12),
                bg="#FFFFFF",
                fg="#333",
                anchor="w",
                width=20,
            ).pack(side="left")

            # Поле ввода для изменения количества
            quantity_entry = Entry(row, font=("Arial", 12), width=10, bg="#F5F5F5")
            quantity_entry.insert(0, str(ingredient[2]))  # Заполняем текущее количество
            quantity_entry.pack(side="left", padx=5)

            # Сохраняем виджет для последующего использования
            self.ingredient_entries[ingredient[0]] = quantity_entry

        # Кнопка сохранения изменений
        Button(
            manage_window,
            text="Сохранить изменения",
            font=("Arial", 12),
            bg="#EEE8CD",
            fg="#333",
            command=lambda: self.save_ingredient_changes(dish),
        ).pack(pady=10)

    def save_ingredient_changes(self, dish):
        """Сохраняет изменения количества ингредиентов и проверяет на наличие нулевых значений."""
        ingredient_updated = False

        # Перебираем все записи ингредиентов
        for ingredient_id, entry in self.ingredient_entries.items():
            try:
                # Получаем новое количество ингредиента
                new_quantity = entry.get().strip()
                if not new_quantity.isdigit():
                    raise ValueError("Введите допустимое количество (целое число).")

                new_quantity = int(new_quantity)

                if new_quantity < 0:  # Проверка на отрицательные значения
                    raise ValueError("Количество не может быть отрицательным.")

                # Обновляем количество ингредиента в базе данных
                execute_query(
                    "UPDATE ingredient SET quantity = ? WHERE ingredient_id = ?",
                    (new_quantity, ingredient_id),
                )
                ingredient_updated = True
            except ValueError as e:
                # Показываем сообщение об ошибке, если введено некорректное значение
                messagebox.showerror("Ошибка", f"Ошибка: {e}")
            except Exception as e:
                # Общий обработчик ошибок для базы данных или других исключений
                messagebox.showerror("Ошибка", f"Ошибка при сохранении данных: {str(e)}")

        if ingredient_updated:
            try:
                # Проверка на нулевые ингредиенты после обновления
                ingredients = fetch_data(
                    "SELECT i.ingredient_name FROM ingredient i "
                    "JOIN dishingredient di ON i.ingredient_id = di.ingredient_id "
                    "WHERE di.dish_id = ? AND i.quantity = 0",
                    (dish[0],),
                )

                if ingredients:
                    missing = [ing[0] for ing in ingredients]
                    self.add_to_stoplist(dish, missing=missing)
                else:
                    self.show_message(f"Изменения для блюда '{dish[1]}' успешно сохранены.")

            except Exception as e:
                # Обработка ошибок при проверке ингредиентов
                messagebox.showerror("Ошибка", f"Ошибка при проверке ингредиентов: {str(e)}")

        # Обновление окна управления ингредиентами
        self.manage_ingredients(dish)

    def mark_missing_and_update_entry(self, ingredient, dish):
        execute_query("UPDATE ingredient SET quantity = 0 WHERE ingredient_id = ?", (ingredient[0],))
        self.add_to_stoplist(dish, missing=[ingredient[1]])

    def restore_dish(self, dish):
        execute_query("DELETE FROM stopped_dishes WHERE dish_id = ?", (dish[0],))
        execute_query(
            "INSERT INTO dish (dish_id, dish_name, price) VALUES (?, ?, ?)",
            (dish[0], dish[1], dish[2]),
        )
        self.show_message(f"Блюдо '{dish[1]}' восстановлено в меню.")
        self.load_menu()
        self.load_stopped_dishes()

    def show_message(self, message):
        msg_window = Toplevel(self.window)
        msg_window.title("Сообщение")
        msg_window.geometry("500x300")
        msg_window.config(bg="#D3D3D3")

        Label(
            msg_window,
            text=message,
            font=("Arial", 11),
            bg="#FFFFFF",
            fg="#333",
        ).pack(expand=True, fill="both", padx=20, pady=20)

        Button(
            msg_window,
            text="ОК",
            font=("Arial", 11),
            bg="#EEE8CD",
            fg="#333",
            command=msg_window.destroy,
        ).pack(pady=10)

# Окно администратора
class AdminWindow:
    def __init__(self, parent):
        self.window = Toplevel(parent)
        self.window.title("Администратор")
        self.window.state("zoomed")
        self.window.config(bg="#D3D3D3")

        # Рамка для меню
        self.menu_frame = Frame(self.window, bg="#ffffff", bd=2, relief="groove")
        self.menu_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        # Рамка для стоп-листа
        self.stopped_frame = Frame(self.window, bg="#ffffff", bd=2, relief="groove", width=250)
        self.stopped_frame.pack(side="left", fill="y", padx=10, pady=10)

        # Кнопка для отчетности по ингредиентам
        self.report_button = Button(
            self.window,
            text="Отчет по ингредиентам",
            font=("Arial", 12),
            bg="#EEE8CD",
            fg="#333",
            command=self.generate_ingredients_report
        )
        self.report_button.pack(padx=10, pady=10)

        # Кнопка для отчетности по блюдам
        self.report_button = Button(
            self.window,
            text="Отчет по блюдам",
            font=("Arial", 12),
            bg="#EEE8CD",
            fg="#333",
            command=self.generate_dish_report
        )

        self.report_button.pack(padx=10, pady=10)
        self.load_menu()
        self.load_stopped_dishes()

    def generate_ingredients_report(self):
        # Логика генерации отчета
        Ingredient_data = fetch_data(
            "SELECT ingredient_id, ingredient_name, quantity FROM Ingredient"
        )

        if not Ingredient_data:
            self.show_message("Нет данных для отчета.")
            return

        # Создаем новый Excel файл
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Ингредиенты"

        # Заголовки
        sheet["A1"] = "ID Ингредиента"
        sheet["B1"] = "Название Ингредиента"
        sheet["C1"] = "Количество"

        # Записываем данные в файл
        for row_idx, ingredient in enumerate(Ingredient_data, start=2):
            sheet.cell(row=row_idx, column=1, value=ingredient[0])  # ingredient_id
            sheet.cell(row=row_idx, column=2, value=ingredient[1])  # ingredient_name
            sheet.cell(row=row_idx, column=3, value=ingredient[2])  # quantity

        # Сохраняем файл
        file_name = "ingredients_report.xlsx"
        wb.save(file_name)

        # Показываем сообщение об успешном сохранении
        self.show_message(f"Отчет по ингредиентам сохранен в файл: {file_name}")

    def show_message(self, message):
        msg_window = Toplevel(self.window)
        msg_window.title("Сообщение")
        msg_window.geometry("500x300+600+300")
        msg_window.config(bg="#D3D3D3")

        Label(
            msg_window,
            text=message,
            font=("Arial", 11),
            bg="#FFFFFF",
            fg="#333",
        ).pack(expand=True, fill="both", padx=20, pady=20)

        Button(
            msg_window,
            text="ОК",
            font=("Arial", 11),
            bg="#EEE8CD",
            fg="#333",
            command=msg_window.destroy,
        ).pack(pady=10)

    def generate_dish_report(self):
        # Логика генерации отчета, замените fetch_data на реальную функцию получения данных.
        Dish_data = self.fetch_data("SELECT dish_id, dish_name, price FROM Dish")

        if not Dish_data:
            self.show_message("Нет данных для отчета.")
            return

        # Создаем новый Excel файл
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Блюда"

        # Заголовки
        sheet["A1"] = "ID Блюда"
        sheet["B1"] = "Название"
        sheet["C1"] = "Цена"

        # Записываем данные в файл
        for row_idx, dish in enumerate(Dish_data, start=2):
            sheet.cell(row=row_idx, column=1, value=dish[0])  # dish_id
            sheet.cell(row=row_idx, column=2, value=dish[1])  # dish_name
            sheet.cell(row=row_idx, column=3, value=dish[2])  # price

        # Сохраняем файл
        files_name = "dish_report.xlsx"
        wb.save(files_name)

        # Показываем сообщение об успешном сохранении
        self.show_message(f"Отчет по блюдам сохранен в файл: {files_name}")

    def show_message(self, message):
        """Отображение сообщения в отдельном окне."""
        msg_window = Toplevel(self.window)
        msg_window.title("Сообщение")
        msg_window.geometry("500x300+600+300")
        msg_window.config(bg="#D3D3D3")

        Label(
            msg_window,
            text=message,
            font=("Arial", 11),
            bg="#FFFFFF",
            fg="#333",
        ).pack(expand=True, fill="both", padx=20, pady=20)

        Button(
            msg_window,
            text="ОК",
            font=("Arial", 11),
            bg="#EEE8CD",
            fg="#333",
            command=msg_window.destroy,
        ).pack(pady=10)

    def fetch_data(self, query):
        """Примерная реализация функции для получения данных из базы данных."""
        try:
            # Здесь подключение к вашей базе данных. Пример для SQLite:
            import sqlite3
            conn = sqlite3.connect('menu.db')
            cursor = conn.cursor()
            cursor.execute(query)
            data = cursor.fetchall()
            conn.close()
            return data
        except Exception as e:
            print(f"Ошибка при запросе данных: {e}")
            return []

    def load_menu(self):
        for widget in self.menu_frame.winfo_children():
            widget.destroy()

        dishes = fetch_data("SELECT d.dish_id, d.dish_name, d.price FROM dish d")

        Label(
            self.menu_frame,
            text="Меню:",
            font=("Arial", 16, "bold"),
            bg="#ffffff",
            fg="#333"
        ).pack(anchor="w", padx=10, pady=5)

        for dish in dishes:
            frame = Frame(self.menu_frame, bg="#ffffff")
            frame.pack(fill="x", padx=10, pady=5)

            Label(
                frame,
                text=f"{dish[1]} - {dish[2]} ₽",
                font=("Arial", 12),
                bg="#ffffff",
                anchor="w",
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Редактировать блюдо",
                font=("Arial", 10),
                bg="#EEE8CD",
                fg="#333",
                command=lambda r=dish: self.edit_dish(r),
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Добавить в стоп-лист",
                font=("Arial", 10),
                bg="#CDC8B1",
                fg="#333",
                command=lambda r=dish: self.add_to_stoplist(r),
            ).pack(side="left", padx=5)

    def load_stopped_dishes(self):
        for widget in self.stopped_frame.winfo_children():
            widget.destroy()

        stopped_dishes = fetch_data("SELECT dish_id, dish_name, price FROM stopped_dishes")

        Label(
            self.stopped_frame,
            text="Стоп-лист:",
            font=("Arial", 16, "bold"),
            bg="#ffffff",
            fg="#333"
        ).pack(anchor="w", padx=10, pady=5)

        for dish in stopped_dishes:
            frame = Frame(self.stopped_frame, bg="#ffffff")
            frame.pack(fill="x", padx=10, pady=5)

            Label(
                frame,
                text=f"{dish[1]} - {dish[2]} ₽",
                font=("Arial", 12),
                bg="#ffffff",
                anchor="w",
            ).pack(side="left", padx=5)

            Button(
                frame,
                text="Вернуть в меню",
                font=("Arial", 10),
                bg="#EEE8CD",
                fg="#333",
                command=lambda r=dish: self.restore_dish(r),
            ).pack(side="left", padx=5)

    def edit_dish(self, dish):
        edit_window = Toplevel(self.window)
        edit_window.title(f"Редактировать блюдо: {dish[1]}")
        edit_window.geometry("400x300")
        edit_window.config(bg="#FFFFFF")

        Label(edit_window, text="Редактирование блюда", font=("Arial", 14), bg="#FFFFFF").pack(pady=10)

        Label(edit_window, text="Название блюда:", font=("Arial", 12), bg="#FFFFFF").pack(anchor="w", padx=10, pady=5)
        name_entry = Entry(edit_window, font=("Arial", 12))
        name_entry.pack(fill="x", padx=10, pady=5)
        name_entry.insert(0, dish[1])

        Label(edit_window, text="Цена блюда (₽):", font=("Arial", 12), bg="#FFFFFF").pack(anchor="w", padx=10, pady=5)
        price_entry = Entry(edit_window, font=("Arial", 12))
        price_entry.pack(fill="x", padx=10, pady=5)
        price_entry.insert(0, dish[2])

        def save_changes():
            new_name = name_entry.get().strip()
            new_price = price_entry.get().strip()

            if not new_name or not new_price.isdigit():
                self.show_message("Введите корректные данные для имени и цены!")
                return

            execute_query(
                "UPDATE dish SET dish_name = ?, price = ? WHERE dish_id = ?",
                (new_name, int(new_price), dish[0]),
            )
            self.show_message(f"Блюдо '{dish[1]}' обновлено.")
            self.load_menu()
            edit_window.destroy()

        Button(edit_window, text="Сохранить изменения", font=("Arial", 12), bg="#EEE8CD", fg="#333", command=save_changes).pack(pady=10)

    def add_to_stoplist(self, dish):
        execute_query("DELETE FROM dish WHERE dish_id = ?", (dish[0],))
        execute_query(
            "INSERT INTO stopped_dishes (dish_id, dish_name, price) VALUES (?, ?, ?)",
            (dish[0], dish[1], dish[2]),
        )
        self.show_message(f"Блюдо '{dish[1]}' добавлено в стоп-лист.")
        self.load_menu()
        self.load_stopped_dishes()

    def restore_dish(self, dish):
        execute_query("DELETE FROM stopped_dishes WHERE dish_id = ?", (dish[0],))
        execute_query(
            "INSERT INTO dish (dish_id, dish_name, price) VALUES (?, ?, ?)",
            (dish[0], dish[1], dish[2]),
        )
        self.show_message(f"Блюдо '{dish[1]}' возвращено в меню.")
        self.load_stopped_dishes()
        self.load_menu()

    def show_message(self, message):
        msg_window = Toplevel(self.window)
        msg_window.title("Сообщение")
        msg_window.geometry("500x300")
        msg_window.config(bg="#D3D3D3")

        Label(
            msg_window,
            text=message,
            font=("Arial", 11),
            bg="#FFFFFF",
            fg="#333",
        ).pack(expand=True, fill="both", padx=20, pady=20)

        Button(
            msg_window,
            text="ОК",
            font=("Arial", 11),
            bg="#EEE8CD",
            fg="#333",
            command=msg_window.destroy,
        ).pack(pady=10)

class ClientWindow:
    def __init__(self, parent):
        self.window = Toplevel(parent)
        self.window.title("Интерфейс клиента")
        self.window.geometry("673x1200")
        self.window.config(bg="#D3D3D3")

        # Рамка для меню
        self.menu_frame = Frame(self.window, bg="#FFFFFF", bd=2, relief="groove", width=300, height=750)
        self.menu_frame.pack_propagate(False)
        font = ("Arial", 20, "bold")  # Для заголовка
        font = ("Arial", 14)  # Для блюд
        self.menu_frame.place(relx=0.5, rely=0.5, anchor="center")  # Центрирование меню

        # Загрузка меню
        self.load_menu()

    def load_menu(self):
        # Удаление старых виджетов в рамке
        for widget in self.menu_frame.winfo_children():
            widget.destroy()

        # Имитация базы данных (замените на вашу функцию connect_db)
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT d.dish_name, d.price FROM dish d")

        # Заголовок меню
        Label(
            self.menu_frame,
            text="Меню:",
            font=("Arial", 16, "bold"),
            bg="#FFFFFF",
            fg="#333"
        ).pack(pady=10)

        # Отображение списка блюд
        dishes = cursor.fetchall()
        if not dishes:
            Label(
                self.menu_frame,
                text="Меню пока пусто.",
                font=("Arial", 12),
                bg="#FFFFFF",
                fg="#555"
            ).pack(pady=5)
        else:
            for dish in dishes:
                Label(
                    self.menu_frame,
                    text=f"{dish[0]} - {dish[1]} ₽",
                    font=("Arial", 12),
                    bg="#FFFFFF",
                    fg="#555"
                ).pack(pady=5)


        conn.close()

if __name__ == "__main__":
    root = Tk()
    app = MainApp(root)
    root.mainloop()
